# utils.py

import requests
import openpyxl
from .models import CustomerInfo

def download_file(access_token, drive_id, file_id, file_path):
    """下载在线 Excel 文件"""
    try:
        url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{file_id}"
        headers = {"Authorization": f"Bearer {access_token}"}
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        download_url = response.json().get("@microsoft.graph.downloadUrl")
        if not download_url:
            raise ValueError("Failed to get the download URL")
        
        file_response = requests.get(download_url)
        file_response.raise_for_status()
        with open(file_path, "wb") as file:
            file.write(file_response.content)
        print(f"File downloaded successfully and saved to {file_path}")
    except Exception as e:
        print(f"File download failed: {e}")

# utils.py

def update_excel_from_db(file_path):
    """根据数据库中的数据更新 Excel 文件"""
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        # 获取所有 CustomerInfo 数据
        customers = CustomerInfo.objects.all()
        print(f"Total customers found in DB: {customers.count()}")  # 输出客户总数，确保数据库查询无误

        for customer in customers:
            row_found = False
            print(f"Checking for customer: {customer.customer_name}")  # 打印当前正在检查的客户名
            
            # 遍历 Excel 中的行，查找匹配的客户
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):  # min_row=2 假设第一行是表头
                if row[0].value == customer.customer_name:  # 假设 Excel 的第一列是客户名称
                    print(f"Found existing row for {customer.customer_name}")  # 找到匹配的行
                    # 更新现有行
                    row[1].value = customer.functional_requirements
                    row_found = True
                    break

            # 如果没有找到匹配的行，添加新行
            if not row_found:
                print(f"Adding new row for {customer.customer_name}")  # 添加新行的日志
                new_row = [customer.functional_requirements]
                sheet.append(new_row)

        # 保存修改后的文件
        wb.save(file_path)
        print(f"Excel file updated and saved to {file_path}")
    except Exception as e:
        print(f"Error updating Excel file: {e}")


def upload_file(access_token, drive_id, file_id, file_path):
    """将修改后的 Excel 文件上传到 OneDrive"""
    try:
        url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{file_id}/content"
        headers = {
            "Authorization": f"Bearer {access_token}",
            "Content-Type": "application/octet-stream",
        }
        with open(file_path, "rb") as file:
            response = requests.put(url, headers=headers, data=file)

        if response.status_code == 200:
            print("File uploaded successfully")
        else:
            print(f"File upload failed: {response.status_code}, {response.json()}")
    except Exception as e:
        print(f"File upload failed: {e}")
